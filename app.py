from openpyxl import load_workbook
from openpyxl import Workbook

from mypackage.functions import search_name_for_value
from mypackage.functions import search_name_for_value_other
from mypackage.functions import set_column_headers
from mypackage.functions import add_change_values_cells
from mypackage.functions import return_values_cell

from mypackage.app_functions import return_date_value_dict
from mypackage.app_functions import return_users_list
from mypackage.app_functions import split_all_values_list

# from tkinter import Tk
# from tkinter import Button
# from tkinter import filedialog
# from tkinter import Text
# from tkinter import Label

from itertools import zip_longest
import re

NUMS_CLEAR = '0|1|2|3|4|5|6|7|8|9'


def main():
    book_data = load_workbook('zhurnal_medosmotrov_1_12-31_12.xlsx')
    book_write = load_workbook('otchyot_po_ASMO_za_mart_2021.xlsx')
    sheet = book_data['Sheet']

    ''' Данные страниц '''
    sheet_cpp = book_write['ЦПП']
    sheet_galaxy = book_write['Космос']
    sheet_vatutina = book_write['Ватутина']
    sheet_spc_1 = book_write['СПЦ-1']
    sheet_spc_2 = book_write['СПЦ-2']
    sheet_FOeM = book_write['ФОиМ']
    sheet_COI = book_write['ЦОИ']
    sheet_zhdc = book_write['ЖДЦ']
    sheet_atc = book_write['АТЦ']
    sheet_espc = book_write['ЭСПЦ']
    sheet_ambulance = book_write['скорая помощь']

    all_values_list = split_all_values_list(sheet, 'A')
    users_list = return_users_list(all_values_list)
    # admittance_dict = return_date_value_dict(users_list, 'Допуск')

    sheet_galaxy_list = []
    sheet_cpp_list = []
    sheet_vatutina_list = []
    sheet_spc_1_list = []
    sheet_spc_2_list = []
    sheet_FOeM_list = []
    sheet_COI_list = []
    sheet_zhdc_list = []
    sheet_atc_list = []
    sheet_espc_list = []
    sheet_ambulance_list = []

    for user in users_list:
        try:
            if user[1] == '"Терминал МП ""Космос"""':
                sheet_galaxy_list.append(user)
            if user[1] == 'Терминал3 КПБ АТЦ (ЭК)':
                sheet_galaxy_list.append(user)
            if user[1] == 'Терминал1 КПБ АТЦ (ЭК)':
                sheet_galaxy_list.append(user)
            if user[1] == 'Терминал4 КПБ АТЦ (ЭК)':
                sheet_galaxy_list.append(user)
            if user[1] == 'Терминал2 КПБ АТЦ (ЭК)':
                sheet_galaxy_list.append(user)
            if user[1] == '"Терминал МП ""Космос"""':
                sheet_galaxy_list.append(user)
            if user[1] == '"Терминал МП ""Космос"""':
                sheet_galaxy_list.append(user)
            if user[1] == '"Терминал МП ""Космос"""':
                sheet_galaxy_list.append(user)
            if user[1] == '"Терминал МП ""Космос"""':
                sheet_galaxy_list.append(user)
            if user[1] == '"Терминал МП ""Космос"""':
                sheet_galaxy_list.append(user)
        except IndexError:
            pass
    print(sheet_galaxy_list)


if __name__ == '__main__':
    '''workbook = Workbook()
    worksheet = workbook.active
    with open('zhurnal_medosmotrov_1_12-31_12.csv', 'r') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                for idx, val in enumerate(col.split(',')):
                    cell = worksheet.cell(row=r + 1, column=c + 1)
                    cell.value = val
    workbook.save('zhurnal_medosmotrov_1_12-31_12.xlsx')'''
    main()