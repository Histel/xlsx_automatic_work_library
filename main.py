from openpyxl import load_workbook, Workbook
from mypackage.functions import search_name_for_value, search_name_for_value_other, set_column_headers


def main():

    excel_document = load_workbook('test.xlsx')
    new_book = Workbook()

    sheet = excel_document['Sheet1']
    new_sheet = new_book.active

    new_sheet[1][0].value = 'hello world'
    set_column_headers(new_sheet, B='Hello World!', C='Goodbye')

    new_book.save('new_book.xlsx')
    new_book.close()

    turnout = search_name_for_value(sheet, 'B', 'I', 'Москва')
    print(turnout)

    turnout_and_value = search_name_for_value_other(sheet, 'C', 'G', 'I', 'Москва')
    print(turnout_and_value)


if __name__ == '__main__':
    main()